"""
read_excel.py - Lecture diagnostique des fichiers Excel (SANS import BDD)
Sprint 1 - Étape 1.3

Usage :
  python scripts/read_excel.py --dpgf     → analyse le modèle DPGF
  python scripts/read_excel.py --devis    → analyse le devis PSA
  python scripts/read_excel.py --all      → analyse les deux

Sorties :
  logs/analyse_DPGF.log
  logs/analyse_devis.log
"""
import sys
import re
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional
import openpyxl
from loguru import logger

# ── Chemins ────────────────────────────────────────────────────
PROJECT_DIR = Path(__file__).parent.parent
LOG_DIR     = PROJECT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

DPGF_FILE  = PROJECT_DIR / "2_ DPGF_Modèle_ 10-04-2026 (2).xlsx"
DEVIS_FILE = PROJECT_DIR / "PSA_ Urgences_ Devis_ 10-04-2026.xlsx"

# ── Règles de détection du ratio_type ─────────────────────────
UNITS_UNITAIRE  = {"u", "ens", "ensemble", "ft", "forfait"}
UNITS_SURFACIQUE = {"m²", "m2", "m ²"}


# ══════════════════════════════════════════════════════════════
# Data structures
# ══════════════════════════════════════════════════════════════

@dataclass
class ArticleRow:
    row_num:            int
    row_type:           str          # 'chapter' | 'section' | 'article' | 'subtotal'
    code:               Optional[str]
    designation:        str
    unit:               Optional[str]
    quantity_moe:       Optional[float] = None   # DPGF : Q MOE
    quantity_ent:       Optional[float] = None   # DPGF : Q Entreprise
    quantity:           Optional[float] = None   # Devis : Qté
    unit_price:         Optional[float] = None   # PU HT
    total_ht:           Optional[float] = None   # Montant HT
    ratio_type:         str = "SURFACIQUE"
    ratio_type_source:  str = "auto_unit"
    chapter:            str = ""
    section:            str = ""


# ══════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════

def safe_float(value) -> Optional[float]:
    """Convertit une valeur Excel en float, retourne None si impossible."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(" ", "").replace(",", ".").replace("€", "")
        if cleaned in ("", "SO", "-", "N/A"):
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def safe_str(value) -> Optional[str]:
    """Nettoie une valeur Excel en chaîne."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def detect_ratio_type(unit: Optional[str], chapter_ratio_type: str = "SURFACIQUE"):
    """
    Règle de détection automatique du ratio_type.

    Priorité :
    1. Si l'unité est dans UNITS_UNITAIRE  → UNITAIRE  (auto_unit)
    2. Si l'unité est dans UNITS_SURFACIQUE → SURFACIQUE (auto_unit)
    3. Sinon → hérite du type du chapitre  (auto_chapter)

    Returns: (ratio_type, ratio_type_source)
    """
    if unit:
        unit_lower = unit.strip().lower()
        if unit_lower in UNITS_UNITAIRE:
            return "UNITAIRE", "auto_unit"
        if unit_lower in UNITS_SURFACIQUE:
            return "SURFACIQUE", "auto_unit"
    # Héritage du chapitre
    return chapter_ratio_type, "auto_chapter"


def is_subtotal_row(cell_b_value) -> bool:
    """Détecte les lignes sous-total (formule CONCATENATE ou libellé 'Sous-Total')."""
    if cell_b_value is None:
        return False
    s = str(cell_b_value).strip()
    return s.startswith("Sous-Total") or s.startswith("Sous-total") or s.startswith("SOUS-TOTAL")


def is_chapter_row(cell_a_value, cell_b_value) -> bool:
    """Détecte les lignes chapitre : Col A contient un numéro seul (1, 2, 3)
    et Col B est vide OU Col A est vide et la ligne a une désignation de chapitre majuscule."""
    a = safe_str(cell_a_value)
    b = safe_str(cell_b_value)
    if a and b is None:
        # Col A seule → probable chapitre
        return True
    if a and b and re.match(r"^(CHAPITRE|LOT|COURANTS)", b.upper()):
        return True
    return False


# ══════════════════════════════════════════════════════════════
# Parseur DPGF Modèle
# ══════════════════════════════════════════════════════════════

def parse_dpgf(filepath: Path) -> list[ArticleRow]:
    """
    Parse le modèle DPGF (format v2 : 9 colonnes).
    Structure : en-têtes row 5, données à partir de row 6.
    Colonnes : A=Art. | B=Type ratio | C=Nature | D=DESIGNATION | E=U
               F=Q MOE | G=Q Entreprise | H=PU HT | I=Montant HT

    Col B (Type ratio) est EXPLICIT : Surfacique ou Unitaire → ratio_type_source='explicit'
    Col C (Nature)     est EXPLICIT : Titre (section) ou Article
    Chapitres : col A non vide, col B/C/D vides → ligne de chapitre.
    """
    logger.info(f"Ouverture DPGF : {filepath.name}")
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb["DPGF"]

    rows_parsed: list[ArticleRow] = []
    current_chapter        = ""
    current_chapter_num    = ""
    current_section        = ""
    current_chapter_ratio  = "SURFACIQUE"
    row_order              = 0

    stats = {"chapter": 0, "section": 0, "article": 0, "subtotal": 0, "skipped": 0}

    for excel_row_num, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        col_a = row[0] if len(row) > 0 else None  # Art. / chapitre
        col_b = row[1] if len(row) > 1 else None  # Type ratio (Surfacique|Unitaire)
        col_c = row[2] if len(row) > 2 else None  # Nature (Titre|Article)
        col_d = row[3] if len(row) > 3 else None  # DESIGNATION
        col_e = row[4] if len(row) > 4 else None  # U (unité)
        col_f = row[5] if len(row) > 5 else None  # Q MOE
        col_g = row[6] if len(row) > 6 else None  # Q Entreprise
        col_h = row[7] if len(row) > 7 else None  # PU HT
        col_i = row[8] if len(row) > 8 else None  # Montant HT

        code        = safe_str(col_a)
        ratio_raw   = safe_str(col_b)   # "Surfacique" | "Unitaire" | None
        nature      = safe_str(col_c)   # "Titre" | "Article" | None
        designation = safe_str(col_d)
        unit        = safe_str(col_e)

        # ── Chapitre : col A non vide, col D vide ─────────────
        if code and not designation and not ratio_raw:
            current_chapter     = code
            current_section     = ""
            current_chapter_ratio = "SURFACIQUE"
            rows_parsed.append(ArticleRow(
                row_num=excel_row_num, row_type="chapter",
                code=None, designation=current_chapter, unit=None,
                chapter=current_chapter, section="",
                ratio_type=current_chapter_ratio, ratio_type_source="auto_chapter"
            ))
            stats["chapter"] += 1
            row_order += 1
            continue

        # ── Sous-total : désignation contient "Sous-Total" ─────
        if designation and is_subtotal_row(designation):
            rows_parsed.append(ArticleRow(
                row_num=excel_row_num, row_type="subtotal",
                code=None, designation=designation, unit=None,
                total_ht=safe_float(col_i),
                chapter=current_chapter, section=current_section,
                ratio_type=current_chapter_ratio, ratio_type_source="auto_chapter"
            ))
            stats["subtotal"] += 1
            row_order += 1
            continue

        if not designation:
            stats["skipped"] += 1
            continue

        row_order += 1

        # ── Ratio type explicite depuis col B ──────────────────
        if ratio_raw and ratio_raw.lower() == "unitaire":
            ratio_type, ratio_source = "UNITAIRE", "explicit"
        elif ratio_raw and ratio_raw.lower() == "surfacique":
            ratio_type, ratio_source = "SURFACIQUE", "explicit"
        else:
            ratio_type, ratio_source = detect_ratio_type(unit, current_chapter_ratio)

        # ── Titre (section) vs Article ─────────────────────────
        if nature and nature.lower() == "titre":
            current_section = designation
            rows_parsed.append(ArticleRow(
                row_num=excel_row_num, row_type="section",
                code=code, designation=designation, unit=unit,
                chapter=current_chapter, section=current_section,
                ratio_type=ratio_type, ratio_type_source=ratio_source
            ))
            stats["section"] += 1
        else:
            rows_parsed.append(ArticleRow(
                row_num=excel_row_num, row_type="article",
                code=code, designation=designation, unit=unit,
                quantity_moe=safe_float(col_f),
                quantity_ent=safe_float(col_g),
                unit_price=safe_float(col_h),
                total_ht=safe_float(col_i),
                chapter=current_chapter, section=current_section,
                ratio_type=ratio_type, ratio_type_source=ratio_source
            ))
            stats["article"] += 1

    wb.close()
    logger.info(f"Parsing terminé : {stats}")
    return rows_parsed, stats


# ══════════════════════════════════════════════════════════════
# Parseur Devis PSA — détection grille 6 col vs grille DPGF 9 col
# ══════════════════════════════════════════════════════════════


def _detect_devis_column_layout(ws) -> str:
    """'psa6' = A..F historique ; 'dpgf9' = même grille que parse_dpgf (A..I, en-têtes ~ligne 6)."""
    for rnums in ((5, 6), (6, 7)):
        for tup in ws.iter_rows(min_row=rnums[0], max_row=rnums[1], values_only=True):
            if not tup:
                continue
            texts = [(safe_str(x) or "").lower() for x in tup[:9]]
            blob = " ".join(texts)
            if "ratio" in blob and ("nature" in blob or "désignation" in blob or "designation" in blob):
                return "dpgf9"
    max_scan = min(ws.max_row or 7, 80)
    for row in ws.iter_rows(min_row=7, max_row=max_scan, values_only=True):
        if not row or len(row) < 9:
            continue
        b = safe_str(row[1])
        c = safe_str(row[2])
        d = safe_str(row[3])
        if not (b and c and d):
            continue
        if b.lower() in ("surfacique", "unitaire") and c.lower() in ("titre", "article"):
            return "dpgf9"
    return "psa6"


def _parse_devis_dpgf_grid(ws) -> tuple[list[ArticleRow], dict, Optional[float]]:
    """Feuille devis au format DPGF (B=type ratio, C=Nature, D=Désignation, E=U, …)."""
    rows_parsed: list[ArticleRow] = []
    current_chapter = ""
    current_section = ""
    current_chapter_ratio = "SURFACIQUE"
    total_ht_source = None
    stats = {"chapter": 0, "section": 0, "article": 0, "subtotal": 0, "so": 0, "skipped": 0}

    for excel_row_num, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
        col_a = row[0] if len(row) > 0 else None
        col_b = row[1] if len(row) > 1 else None
        col_c = row[2] if len(row) > 2 else None
        col_d = row[3] if len(row) > 3 else None
        col_e = row[4] if len(row) > 4 else None
        col_f = row[5] if len(row) > 5 else None
        col_g = row[6] if len(row) > 6 else None
        col_h = row[7] if len(row) > 7 else None
        col_i = row[8] if len(row) > 8 else None

        code = safe_str(col_a)
        ratio_raw = safe_str(col_b)
        nature = safe_str(col_c)
        designation = safe_str(col_d)
        unit = safe_str(col_e)

        if designation and re.search(r"TOTAL\s*HT", designation.upper()):
            total_ht_source = safe_float(col_i)
            if total_ht_source is not None:
                logger.info(f"Total HT source détecté : {total_ht_source:,.2f} €")
            continue

        if code and not designation and not ratio_raw:
            current_chapter = code
            current_section = ""
            current_chapter_ratio = "SURFACIQUE"
            rows_parsed.append(
                ArticleRow(
                    row_num=excel_row_num,
                    row_type="chapter",
                    code=None,
                    designation=current_chapter,
                    unit=None,
                    chapter=current_chapter,
                    section="",
                    ratio_type=current_chapter_ratio,
                    ratio_type_source="auto_chapter",
                )
            )
            stats["chapter"] += 1
            continue

        if designation and is_subtotal_row(designation):
            rows_parsed.append(
                ArticleRow(
                    row_num=excel_row_num,
                    row_type="subtotal",
                    code=None,
                    designation=designation,
                    unit=None,
                    total_ht=safe_float(col_i),
                    chapter=current_chapter,
                    section=current_section,
                    ratio_type=current_chapter_ratio,
                    ratio_type_source="auto_chapter",
                )
            )
            stats["subtotal"] += 1
            continue

        if not designation:
            stats["skipped"] += 1
            continue

        if ratio_raw and ratio_raw.lower() == "unitaire":
            ratio_type, ratio_source = "UNITAIRE", "explicit"
        elif ratio_raw and ratio_raw.lower() == "surfacique":
            ratio_type, ratio_source = "SURFACIQUE", "explicit"
        else:
            ratio_type, ratio_source = detect_ratio_type(unit, current_chapter_ratio)

        q_moe = safe_float(col_f)
        q_ent = safe_float(col_g)
        qty_devis = q_ent if q_ent not in (None, 0) else q_moe

        if nature and nature.lower() == "titre":
            current_section = designation
            rows_parsed.append(
                ArticleRow(
                    row_num=excel_row_num,
                    row_type="section",
                    code=code,
                    designation=designation,
                    unit=unit,
                    chapter=current_chapter,
                    section=current_section,
                    ratio_type=ratio_type,
                    ratio_type_source=ratio_source,
                )
            )
            stats["section"] += 1
        else:
            rows_parsed.append(
                ArticleRow(
                    row_num=excel_row_num,
                    row_type="article",
                    code=code,
                    designation=designation,
                    unit=unit,
                    quantity_moe=q_moe,
                    quantity_ent=q_ent,
                    quantity=qty_devis,
                    unit_price=safe_float(col_h),
                    total_ht=safe_float(col_i),
                    chapter=current_chapter,
                    section=current_section,
                    ratio_type=ratio_type,
                    ratio_type_source=ratio_source,
                )
            )
            stats["article"] += 1

    return rows_parsed, stats, total_ht_source


def _parse_devis_psa_six_col(ws) -> tuple[list[ArticleRow], dict, Optional[float]]:
    """Grille historique : A=Art. | B=Désignation | C=U | D=Qté | E=PU | F=Total HT."""
    rows_parsed: list[ArticleRow] = []
    current_chapter = ""
    current_section = ""
    current_chapter_ratio = "SURFACIQUE"
    total_ht_source = None
    stats = {"chapter": 0, "section": 0, "article": 0, "subtotal": 0, "so": 0, "skipped": 0}

    for excel_row_num, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
        col_a, col_b, col_c, col_d, col_e, col_f = (
            row[0],
            row[1],
            row[2],
            row[3] if len(row) > 3 else None,
            row[4] if len(row) > 4 else None,
            row[5] if len(row) > 5 else None,
        )

        designation = safe_str(col_b) or safe_str(col_a)
        if not designation:
            stats["skipped"] += 1
            continue

        code = safe_str(col_a)
        unit = safe_str(col_c)
        qty_raw = col_d
        unit_price = safe_float(col_e)
        total_ht = safe_float(col_f)

        if designation and re.search(r"TOTAL\s*HT", designation.upper()):
            total_ht_source = total_ht
            logger.info(f"Total HT source détecté : {total_ht_source:,.2f} €")
            continue

        if is_subtotal_row(col_b):
            rows_parsed.append(
                ArticleRow(
                    row_num=excel_row_num,
                    row_type="subtotal",
                    code=None,
                    designation=designation,
                    unit=None,
                    total_ht=total_ht,
                    chapter=current_chapter,
                    section=current_section,
                    ratio_type=current_chapter_ratio,
                    ratio_type_source="auto_chapter",
                )
            )
            stats["subtotal"] += 1
            continue

        if code and not re.search(r"\.", str(code)) and not unit and not unit_price:
            current_chapter = designation
            current_section = ""
            if unit and unit.strip().lower() in UNITS_SURFACIQUE:
                current_chapter_ratio = "SURFACIQUE"
            else:
                current_chapter_ratio = "SURFACIQUE"

            rows_parsed.append(
                ArticleRow(
                    row_num=excel_row_num,
                    row_type="chapter",
                    code=code,
                    designation=designation,
                    unit=unit,
                    chapter=current_chapter,
                    section="",
                    ratio_type=current_chapter_ratio,
                    ratio_type_source="auto_chapter",
                )
            )
            stats["chapter"] += 1
            continue

        # ── Sous-chapitre PSA (ex. A=« 2.1 », B=libellé, C–F vides) ───────────
        # Sans ce cas, tout est classé « article » car le code contient un « . »
        # et les articles apparaissent à plat sous le chapitre dans le matching.
        qty_probe = safe_float(qty_raw)
        qty_str_nonempty = isinstance(qty_raw, str) and bool(qty_raw.strip())
        if code and designation:
            parts = str(code).strip().split(".")
            if (
                len(parts) == 2
                and all(p.isdigit() for p in parts)
                and not unit
                and unit_price is None
                and (total_ht is None or total_ht == 0)
                and qty_probe is None
                and not qty_str_nonempty
            ):
                current_section = f"{code} — {designation}"
                rows_parsed.append(
                    ArticleRow(
                        row_num=excel_row_num,
                        row_type="section",
                        code=code,
                        designation=designation,
                        unit=None,
                        chapter=current_chapter,
                        section=current_section,
                        ratio_type=current_chapter_ratio,
                        ratio_type_source="auto_chapter",
                    )
                )
                stats["section"] += 1
                continue

        qty_is_so = isinstance(qty_raw, str) and qty_raw.strip().upper() == "SO"
        if qty_is_so:
            rows_parsed.append(
                ArticleRow(
                    row_num=excel_row_num,
                    row_type="so",
                    code=code,
                    designation=designation,
                    unit=unit,
                    quantity=None,
                    unit_price=unit_price,
                    total_ht=0.0,
                    chapter=current_chapter,
                    section=current_section,
                    ratio_type="UNITAIRE",
                    ratio_type_source="auto_unit",
                )
            )
            stats["so"] += 1
            continue

        quantity = safe_float(qty_raw)
        ratio_type, ratio_source = detect_ratio_type(unit, current_chapter_ratio)
        is_article = (
            bool(code and re.search(r"\.", str(code)))
            or bool(unit)
            or bool(unit_price)
            or (total_ht is not None and total_ht > 0)
        )

        if is_article:
            rows_parsed.append(
                ArticleRow(
                    row_num=excel_row_num,
                    row_type="article",
                    code=code,
                    designation=designation,
                    unit=unit,
                    quantity=quantity,
                    unit_price=unit_price,
                    total_ht=total_ht,
                    chapter=current_chapter,
                    section=current_section,
                    ratio_type=ratio_type,
                    ratio_type_source=ratio_source,
                )
            )
            stats["article"] += 1
        else:
            current_section = designation
            rows_parsed.append(
                ArticleRow(
                    row_num=excel_row_num,
                    row_type="section",
                    code=code,
                    designation=designation,
                    unit=unit,
                    chapter=current_chapter,
                    section=current_section,
                    ratio_type=current_chapter_ratio,
                    ratio_type_source="auto_chapter",
                )
            )
            stats["section"] += 1

    return rows_parsed, stats, total_ht_source


def parse_devis(filepath: Path) -> tuple[list[ArticleRow], dict, Optional[float]]:
    """
    Parse le devis (feuille 1).

    - **psa6** : en-têtes ligne 6, données ligne 7+ — A=Art. | B=Désignation | C=U | D=Qté | E=PU | F=Total
    - **dpgf9** : même grille que le modèle DPGF — B=Type ratio | C=Nature | D=Désignation | E=U | F/G=Q | H=PU | I=Total
    """
    logger.info(f"Ouverture Devis : {filepath.name}")
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    logger.info(f"Feuille : '{sheet_name}'")

    layout = _detect_devis_column_layout(ws)
    logger.info(f"Grille devis détectée : {layout}")
    if layout == "dpgf9":
        rows_parsed, stats, total_ht_source = _parse_devis_dpgf_grid(ws)
    else:
        rows_parsed, stats, total_ht_source = _parse_devis_psa_six_col(ws)

    wb.close()
    logger.info(f"Parsing terminé : {stats}")
    return rows_parsed, stats, total_ht_source


# ══════════════════════════════════════════════════════════════
# Rapport diagnostique
# ══════════════════════════════════════════════════════════════

def rapport_dpgf(rows: list[ArticleRow], stats: dict):
    """Affiche un rapport détaillé du parsing DPGF."""
    articles = [r for r in rows if r.row_type == "article"]
    chapters = [r for r in rows if r.row_type == "chapter"]

    logger.info("=" * 60)
    logger.info("RAPPORT DPGF MODELE")
    logger.info("=" * 60)
    logger.info(f"Total lignes parsées : {len(rows)}")
    logger.info(f"  Chapitres   : {stats['chapter']}")
    logger.info(f"  Sections    : {stats['section']}")
    logger.info(f"  Articles    : {stats['article']}")
    logger.info(f"  Sous-totaux : {stats['subtotal']}")
    logger.info(f"  Ignorées    : {stats['skipped']}")

    # Chapitres détectés
    logger.info("\nChapitres :")
    for ch in chapters:
        logger.info(f"  [{ch.code or '?'}] {ch.designation}")

    # Ratio type distribution
    surfacique = sum(1 for r in articles if r.ratio_type == "SURFACIQUE")
    unitaire   = sum(1 for r in articles if r.ratio_type == "UNITAIRE")
    auto_unit  = sum(1 for r in articles if r.ratio_type_source == "auto_unit")
    auto_chap  = sum(1 for r in articles if r.ratio_type_source == "auto_chapter")
    logger.info(f"\nRatio types :")
    logger.info(f"  SURFACIQUE : {surfacique} ({surfacique/len(articles)*100:.0f}%)")
    logger.info(f"  UNITAIRE   : {unitaire}   ({unitaire/len(articles)*100:.0f}%)")
    logger.info(f"  Source auto_unit    : {auto_unit}")
    logger.info(f"  Source auto_chapter : {auto_chap}")

    # Unités trouvées
    units = {}
    for r in articles:
        u = r.unit or "(vide)"
        units[u] = units.get(u, 0) + 1
    logger.info("\nUnités rencontrées :")
    for u, count in sorted(units.items(), key=lambda x: -x[1]):
        logger.info(f"  '{u}' : {count} articles")

    # Aperçu des 10 premiers articles
    logger.info("\nAperçu (10 premiers articles) :")
    for r in articles[:10]:
        logger.info(f"  [{r.code}] {r.designation[:50]} | {r.unit} | {r.ratio_type}")

    logger.info("=" * 60)


def rapport_devis(rows: list[ArticleRow], stats: dict, total_ht_source: Optional[float]):
    """Affiche un rapport détaillé + vérification de somme."""
    articles = [r for r in rows if r.row_type == "article"]

    logger.info("=" * 60)
    logger.info("RAPPORT DEVIS PSA URGENCES")
    logger.info("=" * 60)
    logger.info(f"Total lignes parsées : {len(rows)}")
    logger.info(f"  Chapitres   : {stats['chapter']}")
    logger.info(f"  Sections    : {stats['section']}")
    logger.info(f"  Articles    : {stats['article']}")
    logger.info(f"  Sans Objet  : {stats['so']}")
    logger.info(f"  Sous-totaux : {stats['subtotal']}")
    logger.info(f"  Ignorées    : {stats['skipped']}")

    # ── Vérification protocole : somme == Total HT source ──
    somme_articles = sum(r.total_ht for r in articles if r.total_ht is not None)
    logger.info(f"\nVérification somme :")
    logger.info(f"  Somme lignes articles : {somme_articles:>12,.2f} €")
    if total_ht_source:
        logger.info(f"  Total HT source       : {total_ht_source:>12,.2f} €")
        ecart = abs(somme_articles - total_ht_source)
        logger.info(f"  Écart                 : {ecart:>12,.2f} €")
        if ecart <= 0.01:
            logger.success("  ASSERTION : OK - Somme cohérente avec le Total HT source")
        else:
            logger.warning(f"  ASSERTION : ÉCART DÉTECTÉ ({ecart:.2f} €) - À vérifier avant import BDD")
    else:
        logger.warning("  Total HT source non trouvé dans le fichier")

    # Lignes avec prix > 0
    with_price = [r for r in articles if r.unit_price and r.unit_price > 0]
    logger.info(f"\nArticles avec prix renseigné : {len(with_price)} / {len(articles)}")

    # Aperçu 10 premiers articles
    logger.info("\nAperçu (10 premiers articles) :")
    for r in articles[:10]:
        logger.info(
            f"  [{r.code}] {(r.designation or '')[:45]:<45} | "
            f"{r.unit or '?':>4} | {r.quantity or 0:>8.1f} | "
            f"{r.unit_price or 0:>10,.2f} € | {r.ratio_type}"
        )

    logger.info("=" * 60)


# ══════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════

def setup_logger(log_file: Path):
    logger.remove()
    logger.add(
        sys.stdout,
        format="<green>{time:HH:mm:ss}</green> | <level>{level: <8}</level> | {message}",
        colorize=True
    )
    logger.add(
        log_file,
        format="{time:YYYY-MM-DD HH:mm:ss} | {level: <8} | {message}",
        rotation="1 MB",
        retention="30 days",
        encoding="utf-8"
    )
    logger.info(f"Log : {log_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Lecture diagnostique des fichiers Excel")
    parser.add_argument("--dpgf",  action="store_true", help="Analyse le modèle DPGF")
    parser.add_argument("--devis", action="store_true", help="Analyse le devis PSA")
    parser.add_argument("--all",   action="store_true", help="Analyse les deux fichiers")
    args = parser.parse_args()

    if not any([args.dpgf, args.devis, args.all]):
        parser.print_help()
        sys.exit(0)

    if args.dpgf or args.all:
        setup_logger(LOG_DIR / "analyse_DPGF.log")
        if not DPGF_FILE.exists():
            logger.error(f"Fichier introuvable : {DPGF_FILE}")
        else:
            rows, stats = parse_dpgf(DPGF_FILE)
            rapport_dpgf(rows, stats)

    if args.devis or args.all:
        setup_logger(LOG_DIR / "analyse_devis.log")
        if not DEVIS_FILE.exists():
            logger.error(f"Fichier introuvable : {DEVIS_FILE}")
        else:
            rows, stats, total_ht = parse_devis(DEVIS_FILE)
            rapport_devis(rows, stats, total_ht)
