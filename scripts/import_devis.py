"""
import_devis.py - Import d'un devis avec Fuzzy Matching vers le référentiel DPGF
Sprint 3 (v3 - coefficients par lot)

Règles métier :
  - Seules les lignes avec unité sont importées dans devis_lines.
  - Fuzzy Match sur la désignation, filtré par chapitre (critère secondaire).
  - Score >= 80% -> mapping_status = 'auto'
  - Score <  80% -> mapping_status = 'pending' (validation manuelle requise)
  - Trois coefficients distincts par projet :
      coef_cfo : Courants Forts
      coef_cfa : Courants Faibles / SSI
      coef_pv  : Photovoltaïque
  - Règle de neutralisation (IMPERATIVE) :
      prix_normalise = unit_price_ht / coef_lot
      (le lot est déduit automatiquement du chapitre de chaque ligne)

Usage :
  python scripts/import_devis.py
  python scripts/import_devis.py --total-ht-cell F251

Arguments optionnels :
  --total-ht-cell CELLULE   Cellule Excel du Total HT global (ex: F251)
                            Si absent, l'utilisateur est invité à le saisir.
"""
import os
import sys
import re
import sqlite3
import argparse
from collections import defaultdict
from pathlib import Path
from dataclasses import dataclass
from typing import Optional
import openpyxl
from loguru import logger
from rapidfuzz import process, fuzz

PROJECT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_DIR))

from scripts.read_excel import parse_devis, safe_float, DEVIS_FILE as _DEVIS_FILE_DEFAULT, ArticleRow

# Permet à app.py d'injecter le chemin du fichier uploadé via l'environnement
DEVIS_FILE = Path(os.environ.get('DEVIS_FILE_OVERRIDE', str(_DEVIS_FILE_DEFAULT)))
from scripts.scoring import unit_aware_score
from scripts.mapping_knowledge import load_index, lookup_in_index, ensure_table as _ensure_knowledge_table
from init_db import init_database

LOG_DIR = PROJECT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

FUZZY_SEUIL_AUTO = 80   # score minimum pour mapping automatique

# Paliers de complexité autorisés
COEF_PALIERS = (1.0, 1.1, 1.2)


def _detect_lot_from_chapter(chapter: str) -> str:
    """Déduit le lot (CFO/CFA/PV) depuis le nom de chapitre du devis.

    Règles :
      'FAIBLE' ou 'SSI' dans le chapitre -> 'CFA'  (Courants Faibles / SSI)
      'PHOTO'  dans le chapitre          -> 'PV'   (Photovoltaïque)
      Tous les autres (dont 'FORT')      -> 'CFO'  (Courants Forts — défaut)
    """
    c = (chapter or "").upper()
    if "FAIBLE" in c or "SSI" in c:
        return "CFA"
    if "PHOTO" in c or "SOLAI" in c:
        return "PV"
    return "CFO"


def _build_context_path(chapter: str, section: str) -> str:
    """Construit le chemin hiérarchique affichable pour une ligne de devis.

    Exemples :
        chapter='COURANTS FORTS', section='DISTRIBUTION'
        → 'COURANTS FORTS > DISTRIBUTION'
        chapter='COURANTS FORTS', section=''
        → 'COURANTS FORTS'
    """
    parts = [p.strip() for p in [chapter, section] if p and p.strip()]
    return " > ".join(parts) if parts else ""


def _compute_reliquats(all_rows: list) -> list:
    """
    Calcule les lignes de reliquat par suivi SÉQUENTIEL des groupes.

    Algorithme : pour chaque sous-total déclaré dans le devis, on compare
    sa valeur à la somme des articles (avec unité) parsés DEPUIS le dernier
    événement de "reset de groupe" (chapitre, section, ou en-tête sans prix).

    Un groupe est délimité par :
      - [début]  : chapitre, section, ou article sans unité ni prix (en-tête de sous-poste)
      - [fin]    : ligne 'subtotal' dans le devis

    Un reliquat est créé quand |sous-total_déclaré - somme_groupe| > 0.01 €.

    Propriétés des lignes reliquat :
      - excel_row_num  = None   (pas de ligne Excel réelle — LIGNE VIRTUELLE)
      - mapping_status = 'unmapped'  (ne s'affiche PAS dans validate_mapping)
      - designation    = "Autres / Reliquat [label_sous-poste]"

    Pour le devis PSA Urgences, tous les groupes sont équilibrés (0 reliquat attendu)
    car les 185 articles avec unité couvrent la totalité du montant.
    """
    reliquats = []

    current_chapter  = ""
    current_section  = ""
    group_label      = ""   # désignation du sous-poste courant
    group_articles   = []   # articles avec unité depuis le dernier reset

    for r in all_rows:

        # ── Nouveau chapitre : reset total ───────────────────────
        if r.row_type == "chapter":
            current_chapter = r.designation
            current_section = ""
            group_label     = r.designation
            group_articles  = []

        # ── Nouvelle section : reset ──────────────────────────────
        elif r.row_type == "section":
            current_section = r.designation
            group_label     = r.designation
            group_articles  = []

        # ── Article avec unité : accumule dans le groupe ──────────
        elif r.row_type in ("article", "so") and r.unit:
            group_articles.append(r)

        # ── Article SANS unité NI prix : en-tête de sous-poste ───
        # (ex : "Installation de chantier", "Réseau de terre", …)
        # Ces lignes marquent le début d'un nouveau sous-poste → reset
        elif r.row_type == "article" and not r.unit and not r.unit_price:
            group_label    = r.designation
            group_articles = []

        # ── Sous-total : ferme le groupe, calcule le reliquat ─────
        elif r.row_type == "subtotal" and r.total_ht is not None:
            declared   = r.total_ht
            actual     = sum(a.total_ht or 0 for a in group_articles)
            reliquat_v = round(declared - actual, 2)

            if abs(reliquat_v) > 0.01:
                label = group_label or current_section or current_chapter or "Divers"
                reliquats.append(ArticleRow(
                    row_num   = None,          # ← LIGNE VIRTUELLE : pas de n° Excel réel
                    row_type  = "article",
                    code      = None,
                    designation = f"Autres / Reliquat {label}",
                    unit      = "ens",
                    quantity  = 1.0,
                    unit_price = reliquat_v,
                    total_ht  = reliquat_v,
                    chapter   = current_chapter,
                    section   = current_section,
                    ratio_type        = "SURFACIQUE",
                    ratio_type_source = "auto_chapter",
                ))
                logger.debug(
                    f"  Reliquat [{label}] : "
                    f"declare={declared:,.2f} - actual={actual:,.2f} = {reliquat_v:+,.2f} EUR"
                )

            # Toujours reset après un sous-total (le groupe est clos)
            group_articles = []
            group_label    = current_section or current_chapter

    return reliquats


# ══════════════════════════════════════════════════════════════
# Lecture cellule Total HT
# ══════════════════════════════════════════════════════════════

# Référence de cellule Excel : une ou plusieurs lettres suivies d'un ou plusieurs
# chiffres (ex: A1, F251, AB1024). Utilisé pour distinguer une saisie "cellule"
# d'une saisie "montant numérique direct".
_CELL_REF_RE = re.compile(r"^[A-Z]+[1-9][0-9]*$")


def read_total_ht_from_cell(filepath: Path, cell_ref: str) -> Optional[float]:
    """
    Lit la valeur du Total HT à contrôler.

    Deux modes de saisie sont acceptés pour `cell_ref` :
      1. Référence de cellule Excel (ex: 'F251') → lue dans la 1ère feuille.
      2. Montant numérique direct (ex: '10270217,6' ou '10270217.6')
         → l'utilisateur a saisi la valeur au lieu de la cellule.

    Retourne None si la saisie est invalide ou la cellule vide / non numérique.
    """
    raw = (cell_ref or "").strip()
    if not raw:
        logger.warning("Total HT : saisie vide — contrôle d'équilibre ignoré.")
        return None

    normalized = raw.upper()

    # Cas 1 — Référence de cellule valide
    if _CELL_REF_RE.match(normalized):
        try:
            wb = openpyxl.load_workbook(str(filepath), data_only=True)
            ws = wb[wb.sheetnames[0]]
            value = ws[normalized].value
            wb.close()
        except (KeyError, ValueError) as e:
            logger.warning(f"Lecture cellule {normalized} impossible : {e}")
            return None
        result = safe_float(value)
        if result is None:
            logger.warning(f"Cellule {normalized} = '{value}' -> non convertible en float")
        else:
            logger.info(f"Total HT lu en {normalized} : {result:,.2f} €")
        return result

    # Cas 2 — Montant numérique direct (tolère virgule FR, espaces, €)
    cleaned = raw.replace("€", "").replace(" ", "").replace("\u00a0", "").replace(",", ".")
    try:
        result = float(cleaned)
        logger.info(f"Total HT saisi directement : {result:,.2f} €")
        return result
    except ValueError:
        logger.warning(
            f"Total HT : '{raw}' n'est ni une cellule Excel valide (ex: F251) "
            "ni un montant numérique. Contrôle d'équilibre ignoré."
        )
        return None


# ══════════════════════════════════════════════════════════════
# Fuzzy Matching
# ══════════════════════════════════════════════════════════════

@dataclass
class MatchResult:
    dpgf_id:       int
    designation:   str
    score:         float
    method:        str    # 'chapter_filtered' | 'global'


def _build_dpgf_index(conn: sqlite3.Connection) -> tuple[list[dict], dict[str, list[dict]], dict[int, dict]]:
    """
    Charge tous les articles DPGF et les indexe par chapitre + par id.

    Returns:
        all_articles   : liste de tous les articles (pour fuzzy)
        by_chapter     : dict {chapter_name: [articles]}
        by_id          : dict {id: article} pour lookup O(1) lors du contrôle de cohérence
    """
    cur = conn.execute(
        "SELECT id, designation, chapter, unit, ratio_type, ratio_type_source "
        "FROM dpgf_articles WHERE row_type = 'article'"
    )
    all_articles = [
        {
            "id": r[0], "designation": r[1], "chapter": r[2],
            "unit": r[3], "ratio_type": r[4], "ratio_type_source": r[5],
        }
        for r in cur.fetchall()
    ]

    by_chapter: dict[str, list[dict]] = {}
    by_id: dict[int, dict] = {}
    for a in all_articles:
        by_chapter.setdefault(a["chapter"], []).append(a)
        by_id[a["id"]] = a

    logger.info(
        f"Index DPGF chargé : {len(all_articles)} articles, "
        f"{len(by_chapter)} chapitres"
    )
    return all_articles, by_chapter, by_id


def _find_closest_chapter(devis_chapter: str, dpgf_chapters: list[str]) -> Optional[str]:
    """
    Trouve le chapitre DPGF le plus proche d'un chapitre devis (fuzzy sur les noms).
    Retourne None si le meilleur score < 40%.
    """
    if not devis_chapter or not dpgf_chapters:
        return None
    result = process.extractOne(devis_chapter, dpgf_chapters, scorer=fuzz.WRatio)
    if result and result[1] >= 40:
        return result[0]
    return None


def fuzzy_match(
    designation: str,
    chapter: str,
    all_articles: list[dict],
    by_chapter: dict[str, list[dict]],
    unit: Optional[str] = None,
    knowledge: Optional[dict] = None,
) -> MatchResult:
    """
    Trouve le meilleur article DPGF pour une désignation de devis.

    Stratégie :
    1. Filtrage par chapitre (fuzzy sur les noms de chapitres).
    2. Fuzzy Match sur la désignation dans le sous-ensemble chapitre.
    3. Si score < seuil minimal (60%) ou pas de candidat chapitre -> recherche globale.
    4. Pénalité unité appliquée sur le score final (×0.5 si familles différentes).
    5. Retourne le meilleur résultat après pénalité.
    """
    # ── Lookup connaissance mémorisée (priorité absolue) ─────────
    if knowledge is not None:
        hit = lookup_in_index(knowledge, designation, unit)
        if hit:
            return MatchResult(
                dpgf_id     = hit["id"],
                designation = hit["designation"],
                score       = 100.0,
                method      = "knowledge",
            )

    dpgf_chapters = list(by_chapter.keys())
    closest_chap  = _find_closest_chapter(chapter, dpgf_chapters)

    best_chapter_match: Optional[MatchResult] = None

    if closest_chap:
        candidates = by_chapter[closest_chap]
        desigs     = [a["designation"] for a in candidates]
        result     = process.extractOne(designation, desigs, scorer=fuzz.WRatio)
        if result:
            art   = candidates[result[2]]
            score = unit_aware_score(result[1], unit, art.get("unit"))
            best_chapter_match = MatchResult(
                dpgf_id     = art["id"],
                designation = result[0],
                score       = score,
                method      = "chapter_filtered",
            )

    # Recherche globale
    all_desigs    = [a["designation"] for a in all_articles]
    global_result = process.extractOne(designation, all_desigs, scorer=fuzz.WRatio)
    if global_result:
        art   = all_articles[global_result[2]]
        score = unit_aware_score(global_result[1], unit, art.get("unit"))
        best_global = MatchResult(
            dpgf_id     = art["id"],
            designation = global_result[0],
            score       = score,
            method      = "global",
        )
    else:
        best_global = None

    # Retenir le meilleur des deux (après pénalité)
    candidates_results = [r for r in [best_chapter_match, best_global] if r is not None]
    if not candidates_results:
        return MatchResult(dpgf_id=-1, designation="", score=0.0, method="none")

    return max(candidates_results, key=lambda r: r.score)


# ══════════════════════════════════════════════════════════════
# Saisie interactive des paramètres projet
# ══════════════════════════════════════════════════════════════

def prompt_project_params(total_ht_cell_arg: Optional[str], args=None) -> dict:
    """
    Interroge l'utilisateur pour les paramètres du projet.
    Si args contient --non-interactive, utilise les valeurs CLI sans saisie clavier.
    Retourne un dict prêt à insérer dans `projects`.
    """
    # ── Mode non-interactif (ré-import automatique) ───────────
    if args and args.non_interactive:
        categories = [
            "Hospitalier", "Education", "Bureaux", "Industrie",
            "Commerce", "Logement", "Autre"
        ]
        cat_idx = (args.category_id or 3) - 1
        if not (0 <= cat_idx < len(categories)):
            logger.error(f"--category-id doit être entre 1 et {len(categories)}")
            sys.exit(1)
        coef_global = args.coef or 1.0  # rétrocompat : --coef applique aux 3 lots
        params = {
            "name":             args.name or "PSA Urgences",
            "devis_date":       args.date or "2024-01-01",
            "category_name":    categories[cat_idx],
            "surface_sdo":      args.sdo or 5000.0,
            "coef_cfo":         args.coef_cfo if args.coef_cfo is not None else coef_global,
            "coef_cfa":         args.coef_cfa if args.coef_cfa is not None else coef_global,
            "coef_pv":          args.coef_pv  if args.coef_pv  is not None else coef_global,
            "puissance_pv_kwp": args.kwp or 0.0,
            "total_ht_cell":    (total_ht_cell_arg or args.total_ht_cell or "F251").upper(),
        }
        logger.info("Mode non-interactif :")
        for k, v in params.items():
            logger.info(f"  {k} = {v}")
        return params

    # ── Mode interactif ───────────────────────────────────────
    print("\n" + "=" * 60)
    print("  IMPORT DEVIS - Parametres du projet")
    print("=" * 60)

    # Nom projet
    name = input("Nom du projet [PSA Urgences] : ").strip() or "PSA Urgences"

    # Date du devis
    while True:
        date_str = input("Date du devis (YYYY-MM-DD) [2024-01-01] : ").strip() or "2024-01-01"
        if re.match(r"^\d{4}-\d{2}-\d{2}$", date_str):
            break
        print("  Format attendu : YYYY-MM-DD")

    # Catégorie
    categories = [
        "Hospitalier", "Education", "Bureaux", "Industrie",
        "Commerce", "Logement", "Autre"
    ]
    print("Categorie :")
    for i, cat in enumerate(categories, 1):
        print(f"  {i}. {cat}")
    while True:
        choice = input("Numero de categorie [3 = Bureaux] : ").strip() or "3"
        if choice.isdigit() and 1 <= int(choice) <= len(categories):
            category_name = categories[int(choice) - 1]
            break
        print(f"  Entrez un numero entre 1 et {len(categories)}")

    # Surface SDO
    while True:
        sdo_str = input("Surface SDO en m2 (obligatoire) : ").strip()
        try:
            sdo = float(sdo_str.replace(",", "."))
            if sdo > 0:
                break
        except ValueError:
            pass
        print("  Valeur numerique > 0 requise.")

    # Coefficients de complexité par lot
    print("Coefficients de complexite par lot (paliers : 1.0 | 1.1 | 1.2) :")
    coef_cfo = coef_cfa = coef_pv = 1.0
    for lot_label, lot_key in [("CFO - Courants Forts", "cfo"),
                                ("CFA - Courants Faibles/SSI", "cfa"),
                                ("PV  - Photovoltaique", "pv")]:
        while True:
            val_str = input(f"  Coef {lot_label} [1.0] : ").strip() or "1.0"
            try:
                val = float(val_str.replace(",", "."))
                if val in COEF_PALIERS:
                    break
            except ValueError:
                pass
            print("  Valeur attendue : 1.0 | 1.1 | 1.2")
        if lot_key == "cfo":
            coef_cfo = val
        elif lot_key == "cfa":
            coef_cfa = val
        else:
            coef_pv = val

    # Puissance PV (optionnelle)
    kwp_str = input("Puissance PV installee en kWp [0] : ").strip() or "0"
    try:
        puissance_pv_kwp = float(kwp_str.replace(",", "."))
    except ValueError:
        puissance_pv_kwp = 0.0

    # Cellule Total HT
    if total_ht_cell_arg:
        cell = total_ht_cell_arg
        print(f"Cellule Total HT : {cell} (passee en argument)")
    else:
        cell = input("Dans quelle cellule se trouve le Total HT global ? [F251] : ").strip() or "F251"

    print("=" * 60 + "\n")

    return {
        "name":             name,
        "devis_date":       date_str,
        "category_name":    category_name,
        "surface_sdo":      sdo,
        "coef_cfo":         coef_cfo,
        "coef_cfa":         coef_cfa,
        "coef_pv":          coef_pv,
        "puissance_pv_kwp": puissance_pv_kwp,
        "total_ht_cell":    cell.upper(),
    }


# ══════════════════════════════════════════════════════════════
# Import principal
# ══════════════════════════════════════════════════════════════

def import_devis(conn: sqlite3.Connection, params: dict) -> int:
    """
    Importe le devis PSA dans projects + devis_lines avec fuzzy mapping.

    Returns : project_id créé.
    """
    # ── Vérification référentiel DPGF ────────────────────────
    n_dpgf = conn.execute("SELECT COUNT(*) FROM dpgf_articles").fetchone()[0]
    if n_dpgf == 0:
        logger.error("dpgf_articles est vide. Lancez d'abord import_dpgf.py.")
        sys.exit(1)

    # ── Lecture Total HT depuis la cellule Excel ──────────────
    total_ht_source = read_total_ht_from_cell(DEVIS_FILE, params["total_ht_cell"])
    if total_ht_source is None:
        logger.warning(
            f"Cellule {params['total_ht_cell']} vide ou non numérique. "
            "L'assertion de contrôle sera ignorée."
        )

    # ── Parsing du devis ──────────────────────────────────────
    if not DEVIS_FILE.exists():
        logger.error(f"Fichier introuvable : {DEVIS_FILE}")
        sys.exit(1)

    rows, stats, total_ht_auto = parse_devis(DEVIS_FILE)
    logger.info(f"Parsing devis : {stats}")

    # Filtrage (2026-04-21) : toutes les lignes article/so avec unit OU total_ht > 0.
    # Les lignes forfaitaires sans unité mais avec un Montant HT doivent apparaître
    # dans le mapping manuel pour que l'utilisateur puisse les corriger.
    def _keep_for_import(r) -> bool:
        if r.row_type not in ("article", "so"):
            return False
        if r.unit:
            return True
        return r.total_ht is not None and r.total_ht > 0

    lines_with_unit    = [r for r in rows if _keep_for_import(r)]
    lines_skipped      = [r for r in rows if r.row_type == "article" and not _keep_for_import(r)]

    logger.info(f"Lignes à importer (unité OU montant HT) : {len(lines_with_unit)}")
    logger.info(f"Lignes ignorées (sans unité ET sans montant) : {len(lines_skipped)}")

    if lines_skipped:
        logger.debug("Désignations ignorées :")
        for r in lines_skipped[:10]:
            logger.debug(f"  {r.designation[:60]}")
        if len(lines_skipped) > 10:
            logger.debug(f"  ... (+{len(lines_skipped) - 10} autres)")

    # ── Récupération category_id ──────────────────────────────
    cat_row = conn.execute(
        "SELECT id FROM building_categories WHERE name = ?",
        (params["category_name"],)
    ).fetchone()
    category_id = cat_row[0] if cat_row else None

    # ── Création du projet ────────────────────────────────────
    cur = conn.execute(
        """
        INSERT INTO projects
            (name, source_file, devis_date, category_id, surface_sdo,
             coef_cfo, coef_cfa, coef_pv, puissance_pv_kwp, total_ht_source)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            params["name"],
            DEVIS_FILE.name,
            params["devis_date"],
            category_id,
            params["surface_sdo"],
            params["coef_cfo"],
            params["coef_cfa"],
            params["coef_pv"],
            params["puissance_pv_kwp"],
            total_ht_source,
        ),
    )
    project_id = cur.lastrowid
    conn.commit()
    logger.info(
        f"Projet cree : id={project_id} | {params['name']} | "
        f"coef CFO={params['coef_cfo']} CFA={params['coef_cfa']} PV={params['coef_pv']}"
    )

    # ── Index DPGF pour fuzzy ─────────────────────────────────
    all_articles, by_chapter, dpgf_by_id = _build_dpgf_index(conn)

    # ── Index des connaissances mémorisées (lookup O(1)) ──────
    _ensure_knowledge_table(conn)
    knowledge_index = load_index(conn)
    n_known = len(knowledge_index)
    if n_known:
        logger.info(f"Connaissances mémorisées chargées : {n_known} règles")

    # ── Coefficients par lot ──────────────────────────────────
    coef_by_lot = {
        "CFO": params["coef_cfo"],
        "CFA": params["coef_cfa"],
        "PV":  params["coef_pv"],
    }

    # ── Mapping & insertion des lignes ────────────────────────
    inserted        = 0
    mapped_auto     = 0
    mapped_pending  = 0

    logger.info("Demarrage du fuzzy matching...")

    for r in lines_with_unit:
        match = fuzzy_match(r.designation or "", r.chapter, all_articles, by_chapter, unit=r.unit, knowledge=knowledge_index)

        # Statut mapping
        if match.score >= FUZZY_SEUIL_AUTO:
            mapping_status = "auto"
            mapped_auto   += 1
        else:
            mapping_status = "pending"
            mapped_pending += 1

        dpgf_article_id = match.dpgf_id if match.dpgf_id != -1 else None

        # ── Contrôle de cohérence "Full Unitaire" ─────────────
        # Si le référentiel déclare explicitement UNITAIRE mais que
        # le devis ne fournit pas de quantité → montant forfaitaire
        # incompatible → on ne peut pas calculer de PU fiable.
        if dpgf_article_id and dpgf_article_id in dpgf_by_id:
            ref = dpgf_by_id[dpgf_article_id]
            if (
                ref["ratio_type"] == "UNITAIRE"
                and ref["ratio_type_source"] == "explicit"
                and (r.quantity is None or r.quantity == 0)
                and r.total_ht and r.total_ht > 0
            ):
                logger.warning(
                    f"[COHERENCE] Ligne Excel {r.row_num} | "
                    f"'{r.designation[:50]}' | "
                    f"Ref DPGF UNITAIRE (explicit) mais devis sans quantite "
                    f"(montant forfaitaire {r.total_ht:,.2f} EUR) — "
                    f"is_stat_valid mis a 0 pour exclure des ratios"
                )
                is_stat_valid_val = 0
            else:
                is_stat_valid_val = 1
        else:
            is_stat_valid_val = 1

        # Neutralisation par lot (REGLE IMPERATIVE)
        lot = _detect_lot_from_chapter(r.chapter)
        coef_lot = coef_by_lot[lot]
        prix_normalise = None
        if r.unit_price and coef_lot > 0:
            prix_normalise = round(r.unit_price / coef_lot, 4)

        # row_type dans devis_lines : 'so' ou 'article'
        dl_row_type = "so" if r.row_type == "so" else "article"

        context_path = _build_context_path(r.chapter, r.section)

        conn.execute(
            """
            INSERT INTO devis_lines
                (project_id, dpgf_article_id,
                 original_code, original_designation, unit, quantity,
                 unit_price_ht, total_ht, prix_normalise,
                 mapping_status, mapping_score, mapping_candidate,
                 row_type, excel_row_num, context_path, sub_chapter_context,
                 lot, is_stat_valid)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                project_id,
                dpgf_article_id,
                r.code,
                r.designation,
                r.unit,
                r.quantity,
                r.unit_price,
                r.total_ht,
                prix_normalise,
                mapping_status,
                round(match.score, 1),
                match.designation,
                dl_row_type,
                r.row_num,
                context_path,
                context_path,
                lot,
                is_stat_valid_val,
            ),
        )
        inserted += 1

    conn.commit()

    # ── Assertion Total HT (PHASE 1 — articles seuls) ─────────
    # Les reliquats ne sont calculés QUE si l'assertion échoue ici.
    # Pour PSA Urgences : les 185 articles suffisent → assertion OK → 0 reliquat.
    somme_articles = conn.execute(
        "SELECT COALESCE(SUM(total_ht), 0) FROM devis_lines "
        "WHERE project_id = ? AND row_type = 'article'",
        (project_id,)
    ).fetchone()[0]

    import_ok = 0
    if total_ht_source:
        ecart = abs(somme_articles - total_ht_source)
        if ecart <= 0.01:
            import_ok = 1
            logger.success(
                f"ASSERTION OK : somme lignes = {somme_articles:,.2f} EUR "
                f"== Total HT source = {total_ht_source:,.2f} EUR"
            )
            logger.info("Reliquats : inutiles (assertion deja equilibree)")
        else:
            # ── PHASE 2 : reliquats pour combler l'ecart ──────────
            # Uniquement si l'assertion echoue — les reliquats
            # sont des LIGNES VIRTUELLES (pas de n° Excel) qui
            # absorbent l'ecart entre sous-totaux declares et
            # articles effectivement detectes.
            logger.warning(
                f"ASSERTION ECART : somme={somme_articles:,.2f} EUR | "
                f"source={total_ht_source:,.2f} EUR | "
                f"ecart={ecart:,.2f} EUR"
            )
            logger.info("Tentative de correction par lignes virtuelles (reliquats)...")
            reliquat_rows = _compute_reliquats(rows)

            for rel in reliquat_rows:
                rel_context = _build_context_path(rel.chapter, rel.section)
                rel_lot     = _detect_lot_from_chapter(rel.chapter)
                rel_coef    = coef_by_lot[rel_lot]
                rel_prix_normalise = (
                    round(rel.unit_price / rel_coef, 4)
                    if rel.unit_price is not None and rel_coef > 0
                    else None
                )
                conn.execute(
                    """
                    INSERT INTO devis_lines
                        (project_id, dpgf_article_id,
                         original_code, original_designation, unit, quantity,
                         unit_price_ht, total_ht, prix_normalise,
                         mapping_status, mapping_score, mapping_candidate,
                         row_type, excel_row_num, context_path, sub_chapter_context,
                         lot, is_stat_valid)
                    VALUES (?, NULL, NULL, ?, ?, ?, ?, ?, ?,
                            'unmapped', 0.0, '[LIGNE VIRTUELLE - reliquat sous-poste]',
                            'article', NULL, ?, ?, ?, 1)
                    """,
                    (
                        project_id,
                        rel.designation, rel.unit, rel.quantity,
                        rel.unit_price,  rel.total_ht, rel_prix_normalise,
                        rel_context,     rel_context, rel_lot,
                    ),
                )
                inserted += 1
                logger.info(
                    f"  [VIRTUEL] Reliquat : [{rel_context}] {rel.total_ht:+,.2f} EUR"
                )

            if reliquat_rows:
                conn.commit()
                # Re-verification apres reliquats
                somme_v2 = conn.execute(
                    "SELECT COALESCE(SUM(total_ht), 0) FROM devis_lines "
                    "WHERE project_id = ? AND row_type = 'article'",
                    (project_id,)
                ).fetchone()[0]
                ecart_v2 = abs(somme_v2 - total_ht_source)
                if ecart_v2 <= 0.01:
                    import_ok = 1
                    logger.success(
                        f"ASSERTION OK apres reliquats : {somme_v2:,.2f} EUR "
                        f"({len(reliquat_rows)} ligne(s) virtuelle(s) ajoutee(s))"
                    )
                else:
                    logger.warning(
                        f"Ecart residuel apres reliquats : {ecart_v2:,.2f} EUR — "
                        f"structure du devis trop complexe pour correction automatique"
                    )
            else:
                logger.warning(
                    f"Ecart non corrige (aucun reliquat applicable) : {ecart:,.2f} EUR"
                )

    # Mise à jour totaux dans le projet
    conn.execute(
        "UPDATE projects SET total_ht_importe = ?, import_ok = ? WHERE id = ?",
        (somme_articles, import_ok, project_id)
    )
    conn.commit()

    # ── Rapport de mapping ────────────────────────────────────
    _rapport_mapping(conn, project_id, inserted, mapped_auto, mapped_pending)

    return project_id


def _rapport_mapping(
    conn: sqlite3.Connection,
    project_id: int,
    inserted: int,
    mapped_auto: int,
    mapped_pending: int,
):
    """Affiche le rapport de mapping post-import."""
    logger.info("=" * 60)
    logger.info("RAPPORT DE MAPPING")
    logger.info("=" * 60)
    logger.info(f"Lignes importées       : {inserted}")
    logger.success(f"  Mapping AUTO (>=80%) : {mapped_auto}  ({mapped_auto / inserted * 100:.0f}%)")
    if mapped_pending > 0:
        logger.warning(
            f"  Mapping PENDING (<80%): {mapped_pending}  ({mapped_pending / inserted * 100:.0f}%)"
        )
    else:
        logger.success("  Aucun pending — validation manuelle non requise.")

    # Top 10 lignes à valider (score le plus bas)
    if mapped_pending > 0:
        pending_rows = conn.execute(
            """
            SELECT id, original_designation, unit, mapping_score, mapping_candidate
            FROM devis_lines
            WHERE project_id = ? AND mapping_status = 'pending'
            ORDER BY mapping_score ASC
            LIMIT 10
            """,
            (project_id,)
        ).fetchall()
        logger.warning("Top 10 lignes pending (score le plus bas) :")
        for row in pending_rows:
            logger.warning(
                f"  id={row[0]:>5} | score={row[3]:>5.1f}% | "
                f"{(row[1] or '')[:40]:<40} -> {(row[4] or 'NO MATCH')[:40]}"
            )
        logger.info("-> Lancez validate_mapping.py pour valider ces correspondances.")

    logger.info("=" * 60)


# ══════════════════════════════════════════════════════════════
# Entrée principale
# ══════════════════════════════════════════════════════════════

def setup_logger():
    logger.remove()
    # Colorize uniquement en terminal interactif. Quand le script est lancé par
    # subprocess (app.py → import), stdout est un pipe et les codes ANSI
    # s'affichent en clair dans le rapport HTML : "[32m19:34:09[0m | [1mINFO".
    is_tty = hasattr(sys.stdout, "isatty") and sys.stdout.isatty()
    logger.add(
        sys.stdout,
        format="<green>{time:HH:mm:ss}</green> | <level>{level: <8}</level> | {message}",
        colorize=is_tty,
    )
    logger.add(
        LOG_DIR / "import_devis.log",
        format="{time:YYYY-MM-DD HH:mm:ss} | {level: <8} | {message}",
        rotation="1 MB",
        retention="30 days",
        encoding="utf-8",
    )


if __name__ == "__main__":
    setup_logger()

    parser = argparse.ArgumentParser(description="Import devis avec fuzzy matching DPGF")
    parser.add_argument(
        "--total-ht-cell",
        default=None,
        metavar="CELLULE",
        help="Cellule Excel du Total HT global (ex: F251).",
    )
    # ── Mode non-interactif : tous les params passés en CLI ───
    parser.add_argument("--non-interactive", action="store_true",
                        help="Pas de saisie clavier — utilise les valeurs CI-dessous.")
    parser.add_argument("--name",        default=None, help="Nom du projet")
    parser.add_argument("--date",        default=None, help="Date du devis YYYY-MM-DD")
    parser.add_argument("--category-id", type=int, default=None,
                        help="ID categorie batiment (1=Hospitalier ... 7=Autre)")
    parser.add_argument("--sdo",         type=float, default=None, help="Surface SDO m2")
    parser.add_argument("--coef",        type=float, default=None,
                        help="Coef complexite global (applique aux 3 lots si coef-cfo/cfa/pv absent)")
    parser.add_argument("--coef-cfo",    type=float, default=None, dest="coef_cfo",
                        help="Coef lot Courants Forts (1.0|1.1|1.2)")
    parser.add_argument("--coef-cfa",    type=float, default=None, dest="coef_cfa",
                        help="Coef lot Courants Faibles/SSI (1.0|1.1|1.2)")
    parser.add_argument("--coef-pv",     type=float, default=None, dest="coef_pv",
                        help="Coef lot Photovoltaique (1.0|1.1|1.2)")
    parser.add_argument("--kwp",         type=float, default=None,
                        help="Puissance PV installee en kWp")
    args = parser.parse_args()

    params = prompt_project_params(args.total_ht_cell, args)
    conn   = init_database()
    try:
        project_id = import_devis(conn, params)
        logger.success(f"Import termine. Projet id={project_id}")
    finally:
        conn.close()
