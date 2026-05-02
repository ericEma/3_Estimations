"""
export_excel.py — Export DPGF au format Egis (openpyxl)
Sprint 4 : Estimation Élec

Génère un fichier Excel avec :
  - En-tête Egis (couleurs, logo textuel, date)
  - Hiérarchie chapitre / section / article
  - Colonnes : Désignation | U | Quantité | PU HT | Total HT
  - Sous-totaux par chapitre colorés
  - Ligne Total Général en bas
"""

import os
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
except ImportError:
    raise ImportError("openpyxl requis : pip install openpyxl")


# ─── Palette Egis ─────────────────────────────────────────────────────────────

COLOR_EGIS_BLUE   = "003DA5"   # bleu Egis
COLOR_CHAPTER_BG  = "1A2744"   # fond chapitre (dark navy)
COLOR_SECTION_BG  = "263354"   # fond section
COLOR_ARTICLE_ODD = "F5F7FA"   # ligne article impair
COLOR_ARTICLE_EVEN= "FFFFFF"   # ligne article pair
COLOR_SUBTOTAL_BG = "0A1628"   # sous-total (très sombre)
COLOR_TOTAL_BG    = "003DA5"   # total général (bleu Egis)

COLOR_WHITE       = "FFFFFF"
COLOR_DARK        = "0D1117"
COLOR_MUTED       = "8B949E"

FORMAT_EUR = '#,##0.00 "€"'
FORMAT_QTY = '#,##0.00'


def _side(style='thin', color='CCCCCC'):
    return Side(style=style, color=color)


def _border_thin():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)


def _font(bold=False, color=COLOR_DARK, size=10, italic=False):
    return Font(name='Calibri', bold=bold, color=color, size=size, italic=italic)


def _align(horizontal='left', vertical='center', wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


# ─── Export principal ─────────────────────────────────────────────────────────

def export_dpgf_excel(affaire: dict, tree: list, output_path: str):
    """
    Génère le fichier Excel DPGF Egis.

    affaire : dict avec name, client, surface_sdo, ...
    tree    : arborescence retournée par engine_ratios.get_dpgf_tree_with_ratios()
    output_path : chemin complet du fichier à créer
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DPGF"

    # Largeurs de colonnes
    col_widths = {
        'A': 6,    # N° ligne
        'B': 55,   # Désignation
        'C': 8,    # Unité
        'D': 12,   # Quantité
        'E': 14,   # PU HT
        'F': 16,   # Total HT
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    row = 1

    # ── En-tête ───────────────────────────────────────────────────
    row = _write_header(ws, affaire, row)

    # ── Colonnes titres ───────────────────────────────────────────
    row = _write_col_headers(ws, row)

    # ── Corps DPGF ────────────────────────────────────────────────
    chapter_total_cells = []   # pour le total général

    for chapter in tree:
        chapter_start_row = row + 1
        chapter_subtotal_col_refs = []

        # Chapitre
        row = _write_chapter_row(ws, chapter, row)

        for section in chapter.get('sections', []):
            # Section
            row = _write_section_row(ws, section, row)

            # Articles
            for art in section.get('articles', []):
                if not art.get('is_included', True):
                    continue
                row, total_cell = _write_article_row(ws, art, row)
                if total_cell:
                    chapter_subtotal_col_refs.append(total_cell)

        # Sous-total chapitre
        if chapter_subtotal_col_refs:
            row = _write_chapter_subtotal(ws, chapter, chapter_subtotal_col_refs, row)
            chapter_total_cells.append(f"F{row}")

        row += 1   # ligne vide entre chapitres

    # ── Total général ─────────────────────────────────────────────
    total_grand_row = row
    last_row = _write_grand_total(ws, chapter_total_cells, row, affaire)
    row = last_row + 2

    # ── Récapitulatif provisions ──────────────────────────────────
    _write_provisions_summary(ws, affaire, total_grand_row, row)

    # ── Mise en page ──────────────────────────────────────────────
    ws.print_title_rows = '1:3'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.freeze_panes            = 'A5'

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _write_header(ws, affaire, row):
    """Bloc d'en-tête : logo Egis + infos affaire."""

    # Ligne 1 : EGIS | Titre
    ws.merge_cells(f'A{row}:B{row}')
    ws.merge_cells(f'C{row}:F{row}')
    ws[f'A{row}'] = 'EGIS BRANCHE SUD'
    ws[f'A{row}'].font      = _font(bold=True, color=COLOR_WHITE, size=13)
    ws[f'A{row}'].fill      = _fill(COLOR_EGIS_BLUE)
    ws[f'A{row}'].alignment = _align('left')
    ws[f'C{row}'] = 'DÉTAIL QUANTITATIF ESTIMATIF — INSTALLATIONS ÉLECTRIQUES'
    ws[f'C{row}'].font      = _font(bold=True, color=COLOR_WHITE, size=12)
    ws[f'C{row}'].fill      = _fill(COLOR_EGIS_BLUE)
    ws[f'C{row}'].alignment = _align('center')
    ws.row_dimensions[row].height = 22
    row += 1

    # Ligne 2 : Projet
    ws.merge_cells(f'A{row}:B{row}')
    ws.merge_cells(f'C{row}:F{row}')
    ws[f'A{row}'] = f"Affaire : {affaire.get('name', '—')}"
    ws[f'A{row}'].font      = _font(bold=True, color=COLOR_DARK, size=10)
    ws[f'A{row}'].fill      = _fill('E8EEF8')
    ws[f'A{row}'].alignment = _align('left')

    info = []
    if affaire.get('client'):        info.append(f"Client : {affaire['client']}")
    if affaire.get('surface_sdo'):   info.append(f"SDO : {int(affaire['surface_sdo'])} m²")
    info.append(f"Date : {date.today().isoformat()}")

    ws[f'C{row}'] = '  |  '.join(info)
    ws[f'C{row}'].font      = _font(color=COLOR_MUTED, size=9)
    ws[f'C{row}'].fill      = _fill('E8EEF8')
    ws[f'C{row}'].alignment = _align('right')
    ws.row_dimensions[row].height = 16
    row += 1

    return row


def _write_col_headers(ws, row):
    """Ligne d'en-tête des colonnes."""
    headers = ['N°', 'DÉSIGNATION', 'U', 'QUANTITÉ', 'PU HT (€)', 'TOTAL HT (€)']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font      = _font(bold=True, color=COLOR_WHITE, size=10)
        cell.fill      = _fill('1C2B4A')
        cell.alignment = _align('center')
        cell.border    = _border_thin()
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_chapter_row(ws, chapter, row):
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"  {chapter['designation'].upper()}"
    ws[f'A{row}'].font      = _font(bold=True, color=COLOR_WHITE, size=11)
    ws[f'A{row}'].fill      = _fill(COLOR_CHAPTER_BG)
    ws[f'A{row}'].alignment = _align('left')
    ws.row_dimensions[row].height = 18
    return row + 1


def _write_section_row(ws, section, row):
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"    {section['designation']}"
    ws[f'A{row}'].font      = _font(bold=True, color=COLOR_WHITE, size=10, italic=True)
    ws[f'A{row}'].fill      = _fill(COLOR_SECTION_BG)
    ws[f'A{row}'].alignment = _align('left')
    ws.row_dimensions[row].height = 15
    return row + 1


def _write_article_row(ws, art, row):
    """Retourne (new_row, total_cell_ref) ou (new_row, None) si ligne vide."""
    qty   = art.get('qty') or 0
    pu    = art.get('unit_price') or 0
    total = qty * pu

    if total == 0 and qty == 0:
        return row, None

    bg = COLOR_ARTICLE_ODD if row % 2 else COLOR_ARTICLE_EVEN

    # N° (vide pour l'instant)
    ws[f'A{row}'] = ''
    ws[f'A{row}'].fill = _fill(bg)

    # Désignation
    ws[f'B{row}'] = f"      {art.get('designation', '')}"
    ws[f'B{row}'].font      = _font(size=9)
    ws[f'B{row}'].fill      = _fill(bg)
    ws[f'B{row}'].alignment = _align('left', wrap=True)
    ws.row_dimensions[row].height = 14

    # Unité
    ws[f'C{row}'] = art.get('unit', '')
    ws[f'C{row}'].font      = _font(size=9, color=COLOR_MUTED)
    ws[f'C{row}'].fill      = _fill(bg)
    ws[f'C{row}'].alignment = _align('center')

    # Quantité
    ws[f'D{row}'] = qty if qty else None
    ws[f'D{row}'].font           = _font(size=9)
    ws[f'D{row}'].fill           = _fill(bg)
    ws[f'D{row}'].alignment      = _align('right')
    ws[f'D{row}'].number_format  = FORMAT_QTY

    # PU HT
    ws[f'E{row}'] = pu if pu else None
    ws[f'E{row}'].font           = _font(size=9)
    ws[f'E{row}'].fill           = _fill(bg)
    ws[f'E{row}'].alignment      = _align('right')
    ws[f'E{row}'].number_format  = FORMAT_EUR

    # Total HT — formule =D*E
    ws[f'F{row}'] = f'=D{row}*E{row}'
    ws[f'F{row}'].font           = _font(size=9, bold=True)
    ws[f'F{row}'].fill           = _fill(bg)
    ws[f'F{row}'].alignment      = _align('right')
    ws[f'F{row}'].number_format  = FORMAT_EUR

    total_cell = f'F{row}'
    return row + 1, total_cell


def _write_chapter_subtotal(ws, chapter, total_cells, row):
    """Ligne de sous-total chapitre."""
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = f"  Sous-total {chapter['designation']}"
    ws[f'A{row}'].font      = _font(bold=True, color=COLOR_WHITE, size=10)
    ws[f'A{row}'].fill      = _fill(COLOR_SUBTOTAL_BG)
    ws[f'A{row}'].alignment = _align('right')

    formula = '=' + '+'.join(total_cells)
    ws[f'F{row}']              = formula
    ws[f'F{row}'].font         = _font(bold=True, color=COLOR_WHITE, size=11)
    ws[f'F{row}'].fill         = _fill(COLOR_SUBTOTAL_BG)
    ws[f'F{row}'].alignment    = _align('right')
    ws[f'F{row}'].number_format= FORMAT_EUR
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_grand_total(ws, chapter_total_cells, row, affaire):
    """Ligne Total Général. Retourne le dernier numéro de ligne écrit."""
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'TOTAL GÉNÉRAL HT'
    ws[f'A{row}'].font      = _font(bold=True, color=COLOR_WHITE, size=12)
    ws[f'A{row}'].fill      = _fill(COLOR_TOTAL_BG)
    ws[f'A{row}'].alignment = _align('right')

    if chapter_total_cells:
        formula = '=' + '+'.join(chapter_total_cells)
    else:
        formula = 0

    ws[f'F{row}']              = formula
    ws[f'F{row}'].font         = _font(bold=True, color=COLOR_WHITE, size=13)
    ws[f'F{row}'].fill         = _fill(COLOR_TOTAL_BG)
    ws[f'F{row}'].alignment    = _align('right')
    ws[f'F{row}'].number_format= FORMAT_EUR
    ws.row_dimensions[row].height = 20

    # Ligne €/m²
    sdo = affaire.get('surface_sdo', 0)
    if sdo and len(chapter_total_cells) > 0:
        r2 = row + 1
        ws.merge_cells(f'A{r2}:E{r2}')
        ws[f'A{r2}'] = f"  Ratio €/m² SDO ({int(sdo)} m²)"
        ws[f'A{r2}'].font      = _font(color=COLOR_MUTED, size=9, italic=True)
        ws[f'A{r2}'].alignment = _align('right')
        ws[f'F{r2}']              = f'=F{row}/{sdo}'
        ws[f'F{r2}'].font         = _font(color=COLOR_MUTED, size=9, italic=True)
        ws[f'F{r2}'].alignment    = _align('right')
        ws[f'F{r2}'].number_format= FORMAT_EUR
        return r2

    return row


def _write_provisions_summary(ws, affaire, total_row, row):
    """Bloc récapitulatif provisions : Total Sec / 3 taxes / Total Majoré."""
    taux_phase  = float(affaire.get('taux_phase')       or 0)
    taux_incert = float(affaire.get('taux_incertitude') or 0)
    coef_risque = float(affaire.get('coef_risque')      or 0)
    phase_label = affaire.get('phase_etude') or 'APD'

    COLOR_PROV_HEADER = "0D1E2E"
    COLOR_PROV_ROW    = "0A1628"
    COLOR_PROV_RESULT = "142840"

    # ── Titre du bloc ─────────────────────────────────────────────
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = '  RÉCAPITULATIF DES PROVISIONS'
    ws[f'A{row}'].font      = _font(bold=True, color='D5F311', size=10)
    ws[f'A{row}'].fill      = _fill(COLOR_PROV_HEADER)
    ws[f'A{row}'].alignment = _align('left')
    ws.row_dimensions[row].height = 15
    row += 1

    def _prov_line(label, value, fmt=None):
        nonlocal row
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = f'    {label}'
        ws[f'A{row}'].font      = _font(color='AABBCC', size=9)
        ws[f'A{row}'].fill      = _fill(COLOR_PROV_ROW)
        ws[f'A{row}'].alignment = _align('right')
        ws[f'F{row}'] = value
        ws[f'F{row}'].font      = _font(bold=True, color=COLOR_WHITE, size=9)
        ws[f'F{row}'].fill      = _fill(COLOR_PROV_ROW)
        ws[f'F{row}'].alignment = _align('right')
        if fmt:
            ws[f'F{row}'].number_format = fmt
        ws.row_dimensions[row].height = 13
        row += 1

    _prov_line('Total HT Sec',                          f'=F{total_row}',    FORMAT_EUR)
    _prov_line(f'① Taxe Phase ({phase_label})',         taux_phase  / 100,   '0.0%')
    _prov_line('② Taxe Incertitude',                    taux_incert / 100,   '0.0%')
    _prov_line('③ Taxe Risque / Aléa',                  coef_risque / 100,   '0.0%')

    # ── Ligne Total Majoré ─────────────────────────────────────────
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = '  TOTAL HT MAJORÉ  (Budget Projet)'
    ws[f'A{row}'].font      = _font(bold=True, color='D5F311', size=11)
    ws[f'A{row}'].fill      = _fill(COLOR_PROV_RESULT)
    ws[f'A{row}'].alignment = _align('right')

    fp = 1 + taux_phase  / 100
    fi = 1 + taux_incert / 100
    fr = 1 + coef_risque / 100
    ws[f'F{row}'] = f'=F{total_row}*{fp:.6f}*{fi:.6f}*{fr:.6f}'
    ws[f'F{row}'].font      = _font(bold=True, color='D5F311', size=13)
    ws[f'F{row}'].fill      = _fill(COLOR_PROV_RESULT)
    ws[f'F{row}'].alignment = _align('right')
    ws[f'F{row}'].number_format = FORMAT_EUR
    ws.row_dimensions[row].height = 20
