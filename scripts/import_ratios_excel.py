#!/usr/bin/env python3
"""
Bootstrap estimation_ratios.db depuis l'onglet REX du fichier Excel archive.

Onglet 1 (Actualisations) : ignoré — actualisation via moteur app (3 %/an).
Onglet 2 (REX) : une ligne = une affaire historique.
Colonnes O–R (ratios actualisés Excel) : ignorées — recalcul moteur.

Usage :
  python scripts/import_ratios_excel.py
  python scripts/import_ratios_excel.py --file "Ratios/0_ Ratios de prix_ 17-04-2025.xlsx"
  python scripts/import_ratios_excel.py --dry-run
  python scripts/import_ratios_excel.py --replace
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

PROJECT_DIR = Path(__file__).resolve().parent.parent
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from openpyxl import load_workbook

import engine_ratio_referentiel as err

DEFAULT_XLSX = PROJECT_DIR / "Ratios" / "0_ Ratios de prix_ 17-04-2025.xlsx"
DEFAULT_SHEET = "REX"
DATA_START_ROW = 4
HEADER_ROW = 3

VALID_CATEGORIES = {
    "Aéroport", "Bureaux", "Château", "Groupe scolaire", "Collège",
    "EHPAD", "Gymnase", "Hôpital", "Hôtel", "Industrie",
    "Laboratoire", "Logements", "Lycée", "Parking", "Stade",
}

CATEGORY_ALIASES = {
    "aeroport": "Aéroport",
    "bureaux": "Bureaux",
    "chateaux": "Château",
    "chateau": "Château",
    "groupe scolaire": "Groupe scolaire",
    "college": "Collège",
    "ehpad": "EHPAD",
    "gymnase": "Gymnase",
    "hopital": "Hôpital",
    "hotel": "Hôtel",
    "industrie": "Industrie",
    "laboratoire": "Laboratoire",
    "logements": "Logements",
    "lycee": "Lycée",
    "parking": "Parking",
    "stade": "Stade",
}


def _strip_accents(s: str) -> str:
    nf = unicodedata.normalize("NFD", s)
    return "".join(c for c in nf if unicodedata.category(c) != "Mn")


def normalize_category(raw: str | None) -> str | None:
    if not raw:
        return None
    s = str(raw).strip()
    if not s or s.lower() in ("a saisir", "type de batiment"):
        return None
    key = _strip_accents(s).lower()
    if key in CATEGORY_ALIASES:
        return CATEGORY_ALIASES[key]
    for cat in VALID_CATEGORIES:
        if _strip_accents(cat).lower() == key:
            return cat
    return s


def _float_cell(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _year_to_date(year_val) -> str | None:
    if year_val is None or year_val == "":
        return None
    try:
        y = int(float(year_val))
    except (TypeError, ValueError):
        return None
    if y < 1990 or y > datetime.now().year + 1:
        return None
    return f"{y}-06-01"


def parse_rex_row(ws, row_idx: int) -> dict | None:
    name = ws.cell(row_idx, 1).value
    if name is None:
        return None
    name = str(name).strip()
    if not name or name.lower() == "affaires":
        return None

    category = normalize_category(ws.cell(row_idx, 2).value)
    sdo = _float_cell(ws.cell(row_idx, 3).value)
    kwc = _float_cell(ws.cell(row_idx, 4).value)
    devis_date = _year_to_date(ws.cell(row_idx, 5).value)
    notes_val = ws.cell(row_idx, 6).value
    notes = str(notes_val).strip() if notes_val else None

    cfo = _float_cell(ws.cell(row_idx, 7).value)
    cfa = _float_cell(ws.cell(row_idx, 8).value)
    combined = _float_cell(ws.cell(row_idx, 9).value)
    pv = _float_cell(ws.cell(row_idx, 10).value)

    if sdo <= 0:
        return None
    if not category:
        return {"skip": True, "name": name, "reason": "typologie inconnue"}
    if not devis_date:
        return {"skip": True, "name": name, "reason": "année projet absente"}
    if combined <= 0 and (cfo + cfa + pv) <= 0:
        return {"skip": True, "name": name, "reason": "montants HT absents"}

    return {
        "name": name,
        "category_name": category,
        "devis_date": devis_date,
        "surface_sdo": sdo,
        "puissance_pv_kwc": kwc,
        "total_ht_cfo": cfo,
        "total_ht_cfa": cfa,
        "total_ht_pv": pv,
        "total_ht_combined": combined if combined > 0 else None,
        "notes": notes,
    }


def import_rex_file(
    xlsx_path: Path,
    *,
    sheet_name: str = DEFAULT_SHEET,
    batch_id: str | None = None,
    dry_run: bool = False,
    replace: bool = False,
) -> dict:
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"Fichier introuvable : {xlsx_path}")

    batch_id = batch_id or f"excel_rex_{xlsx_path.stem[:40]}"
    batch_id = re.sub(r"[^\w\-]+", "_", batch_id)

    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Onglet « {sheet_name} » absent — dispo : {wb.sheetnames}")
    ws = wb[sheet_name]

    stats = {"imported": 0, "skipped": 0, "errors": 0, "details": []}

    if replace and not dry_run:
        deleted = err.delete_sources_by_batch(batch_id)
        stats["replaced"] = deleted

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        parsed = parse_rex_row(ws, row_idx)
        if parsed is None:
            continue
        if parsed.get("skip"):
            stats["skipped"] += 1
            stats["details"].append(f"SKIP {parsed['name']} — {parsed['reason']}")
            continue

        if dry_run:
            stats["imported"] += 1
            stats["details"].append(
                f"OK {parsed['name']} | {parsed['category_name']} | "
                f"{parsed['devis_date']} | SDO {parsed['surface_sdo']:.0f}"
            )
            continue

        try:
            sid = err.insert_archive_source(
                name=parsed["name"],
                category_name=parsed["category_name"],
                devis_date=parsed["devis_date"],
                surface_sdo=parsed["surface_sdo"],
                puissance_pv_kwc=parsed["puissance_pv_kwc"],
                total_ht_cfo=parsed["total_ht_cfo"],
                total_ht_cfa=parsed["total_ht_cfa"],
                total_ht_pv=parsed["total_ht_pv"],
                total_ht_combined=parsed["total_ht_combined"],
                notes=parsed.get("notes"),
                source_file=xlsx_path.name,
                import_batch_id=batch_id,
            )
            stats["imported"] += 1
            stats["details"].append(f"#{sid} {parsed['name']}")
        except Exception as exc:
            stats["errors"] += 1
            stats["details"].append(f"ERR {parsed['name']} — {exc}")

    wb.close()
    stats["batch_id"] = batch_id
    return stats


def main() -> int:
    parser = argparse.ArgumentParser(description="Import bootstrap ratios Excel (onglet REX)")
    parser.add_argument(
        "--file",
        type=Path,
        default=DEFAULT_XLSX,
        help="Chemin du fichier Excel archive",
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Nom de l'onglet (défaut: REX)")
    parser.add_argument("--batch-id", default=None, help="Identifiant de lot import")
    parser.add_argument("--dry-run", action="store_true", help="Simulation sans écriture BDD")
    parser.add_argument("--replace", action="store_true", help="Supprimer le lot existant avant import")
    args = parser.parse_args()

    try:
        stats = import_rex_file(
            args.file.resolve(),
            sheet_name=args.sheet,
            batch_id=args.batch_id,
            dry_run=args.dry_run,
            replace=args.replace,
        )
    except Exception as exc:
        print(f"Échec : {exc}", file=sys.stderr)
        return 1

    mode = "DRY-RUN" if args.dry_run else "IMPORT"
    print(f"[{mode}] batch={stats.get('batch_id')}")
    print(f"  importés : {stats['imported']}")
    print(f"  ignorés  : {stats['skipped']}")
    print(f"  erreurs  : {stats['errors']}")
    if stats.get("replaced"):
        print(f"  remplacés: {stats['replaced']}")
    for line in stats["details"][:30]:
        print(f"  · {line}")
    if len(stats["details"]) > 30:
        print(f"  … ({len(stats['details']) - 30} lignes supplémentaires)")

    return 0 if stats["errors"] == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
